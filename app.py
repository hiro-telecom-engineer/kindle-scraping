import logging
import time

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
import re

import datetime
import openpyxl

logging.basicConfig(level=logging.INFO)


# 認証の情報
AMAZON_EMAIL = "任意のメールアドレス"
AMAZON_PASSWORD = "任意のパスワード"

chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option(	"excludeSwitches", ['enable-automation'])
driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)


def main():
	excel = make_excel()
	log_in()
	get_book_inf( excel )
	get_comic_inf( excel )
	excel.save_file()
	driver.quit()  # ブラウザーを終了する。

# ログイン処理
def log_in():
	# 注文履歴のページを開く。
	logging.info('Navigating...')
	driver.get("https://read.amazon.co.jp/kindle-library?tabView=all&seriesSortType=acquisition_desc&resourceType=EBOOK&sortType=title#library")
	time.sleep(1)
	# サインインページにリダイレクトされていることを確認する。
	assert 'Amazonサインイン' in driver.title
	email_input = driver.find_element_by_name('email')
	email_input.send_keys(AMAZON_EMAIL)  # name="email" という入力ボックスを埋める。
	email_input.send_keys(Keys.RETURN)
	time.sleep(1)
	password_input = driver.find_element_by_name('password')
	password_input.send_keys(AMAZON_PASSWORD)  # name="password" という入力ボックスを埋める。
	time.sleep(5)
	# フォームを送信する。
	logging.info('Signing in...')
	password_input.send_keys(Keys.RETURN)
	time.sleep(1)


# ページ内を下までスクロール
def scroll():
	while 1:
		html01=driver.page_source
		driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
		time.sleep(3)
		html02=driver.page_source
		if html01!=html02:
			html01=html02
		else:
			break


# コミック以外情報収集
def get_book_inf( excel ):
	scroll()
	book_dict = { "種別":"本", "タイトル":"取得不可", "著者":"取得不可","掲載誌":"-" ,"既刊の巻数":"-", "購入済みの巻数":"-","購入済み":"-","未購入":"-",}
	# 各タイトルごとの著者名取得
	element_title = driver.find_elements_by_css_selector('._2czmS0An9GDlVR9xgpCNOC')
	element_author = driver.find_elements_by_css_selector('._33h88ogkqT8qrfT1uutvBI')
	for (e_title , e_author) in zip( element_title , element_author ):
		book_dict["タイトル"] = e_title.text
		book_dict["著者"] = e_author.text
		excel.add_inf(book_dict)


# コミック情報収集
def get_comic_inf( excel ):
	driver.get('https://read.amazon.co.jp/kindle-library?originType=COMICS&tabView=series&seriesSortType=acquisition_desc&itemView=row')
	time.sleep(1)
	scroll()
	author_inf = []
	author_dict = {"タイトル":"", "著者":""}

	# 各タイトルごとの著者名取得
	element_title = driver.find_elements_by_css_selector('._2czmS0An9GDlVR9xgpCNOC')
	element_author = driver.find_elements_by_css_selector('._33h88ogkqT8qrfT1uutvBI')
	for (e_title , e_author) in zip( element_title , element_author ):
		author_dict["タイトル"] = e_title.text
		author_dict["著者"] = e_author.text
		author_inf.append(author_dict.copy())

	# 各タイトルごとのURL取得
	elem_url=[]
	elems = driver.find_elements_by_tag_name("a")
	for elem in elems:
		book_url = elem.get_attribute("href")
		if "collection" in str(book_url):
			elem_url.append(book_url)

	# 各タイトルごとの情報取得
	for i , book_url in enumerate(elem_url):
		book_dict = { "種別":"コミック", "タイトル":"取得不可", "著者":"取得不可","掲載誌":"取得不可" ,"既刊の巻数":"取得不可", "購入済みの巻数":"取得不可","購入済み":"取得不可","未購入":"取得不可",}
		book_list = []
		driver.get(book_url)
		time.sleep(4)

		# タイトル取得
		book_dict["タイトル"] = author_inf[i]["タイトル"]

		# 著者取得
		book_dict["著者"] = author_inf[i]["著者"]

		# 既刊の巻数、購入済みの巻数取得
		element = driver.find_elements_by_id("series-items-count-text")[0].text
		book_num = re.split(r'このシリーズ（全 | 巻）のうち | 冊を持っています。', element)
		book_dict["既刊の巻数"] = book_num[1]
		book_dict["購入済みの巻数"] = book_num[2]

		# 掲載誌取得と購入済みリスト作成
		element_title = driver.find_elements_by_css_selector('._2czmS0An9GDlVR9xgpCNOC')
		element_author = driver.find_elements_by_css_selector('._33h88ogkqT8qrfT1uutvBI')
		for (e_title , e_author) in zip( element_title , element_author ):
			if re.sub('[^ぁ-んァ-ン一-龥a-zA-Z0-9０-９]', '', book_dict["タイトル"]) != re.sub('[^ぁ-んァ-ン一-龥a-zA-Z0-9０-９]', '', e_title.text):
				if re.sub('[^ぁ-んァ-ン一-龥a-zA-Z0-9０-９]', '', book_dict["タイトル"]) in re.sub('[^ぁ-んァ-ン一-龥a-zA-Z0-9０-９]', '', e_title.text):
					if re.search(r'[0-9０-９]', e_title.text):
						book_list.append(e_title.text.replace('（',' ').replace('）',' ').replace('(',' ').replace(')',' ').replace('巻',' ').replace('　',' ').replace('  ',' ').replace('  ',' ').split(' ')[-4:])
		if 1 > len(book_list):
			print("詳細情報取得不可：" + book_dict["タイトル"])
		else:
			book_dict["掲載誌"] = book_list[0][-2]

			# 購入済み取得
			bought_num = ""
			for book in book_list:
				if 4 > len(book):
					bought_num = bought_num + "1、"
				else:
					if re.compile(r'[0-9０-９]').match(book[-3]):
						bought_num = bought_num + str(int(book[-3])) + "、"
					elif re.compile(r'[0-9０-９]').match(book[-4]):
						bought_num = bought_num + str(int(book[-4])) + "、"
			book_dict["購入済み"] = bought_num[:-1]

			# 未購入取得
			not_purchased_num = ""
			for ctr in range(int(book_dict["既刊の巻数"])):
				add_num = str(ctr+1) + "、"
				for book in book_list:
					if 4 > len(book):
						if ctr + 1 == 1:
							add_num = ""
					else:
						if re.compile(r'[0-9０-９]').match(book[-3]):
							if ctr + 1 == int(book[-3]):
								add_num = ""
								break
						elif re.compile(r'[0-9０-９]').match(book[-4]):
							if ctr + 1 == int(book[-4]):
								add_num = ""
								break
				not_purchased_num = not_purchased_num + add_num
			if "" == not_purchased_num:
				book_dict["未購入"] = "-"
			else:
				book_dict["未購入"] = not_purchased_num[:-1]
		excel.add_inf(book_dict)
	return


# エクセル転記
class make_excel():
	fruits = [ ]
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = '購入履歴一覧'
	write_line = 1
	def __init__( self ) :
		self.ws["A" + str(self.write_line)].value = "種別"
		self.ws["B" + str(self.write_line)].value = "タイトル"
		self.ws["C" + str(self.write_line)].value = "著者"
		self.ws["D" + str(self.write_line)].value = "掲載誌"
		self.ws["E" + str(self.write_line)].value = "既刊の巻数"
		self.ws["F" + str(self.write_line)].value = "購入済みの巻数"
		self.ws["G" + str(self.write_line)].value = "購入済み"
		self.ws["H" + str(self.write_line)].value = "未購入"
		self.write_line += 1
		return

	def add_inf( self , book_dict ) :
		self.ws["A" + str(self.write_line)].value = book_dict["種別"]
		self.ws["B" + str(self.write_line)].value = book_dict["タイトル"]
		self.ws["C" + str(self.write_line)].value = book_dict["著者"]
		self.ws["D" + str(self.write_line)].value = book_dict["掲載誌"]
		self.ws["E" + str(self.write_line)].value = book_dict["既刊の巻数"]
		self.ws["F" + str(self.write_line)].value = book_dict["購入済みの巻数"]
		self.ws["G" + str(self.write_line)].value = book_dict["購入済み"]
		self.ws["H" + str(self.write_line)].value = book_dict["未購入"]
		self.write_line += 1
		print(book_dict)

	def save_file( self ):
		# ファイル保存
		now = datetime.datetime.now()
		file_name = 'kindle情報一覧_{}.xlsx'.format(now.strftime('%Y%m%d_%H%M%S'))
		self.wb.save(file_name)

if __name__ == '__main__':
	main()

